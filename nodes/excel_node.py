import os
import re
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Colours
HOT_FILL  = PatternFill("solid", fgColor="FF4C4C")   # red
WARM_FILL = PatternFill("solid", fgColor="FFD966")   # yellow
HEADER_FILL = PatternFill("solid", fgColor="1F2D3D") # dark navy
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT  = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNS = [
    ("Lead Status",          15),
    ("Business Name",        30),
    ("Category",             20),
    ("Total Reviews",        14),
    ("Phone Number",         18),
    ("Website",              30),
    ("Address",              35),
    ("Star Rating",          12),
    ("Customers Praise",     55),
    ("Customers Complaints", 55),
]


def excel_node(state: dict) -> dict:
    businesses = state["analyzed_businesses"]
    industry   = state["industry"]
    city       = state["city"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # ── Title row ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"Lead Report — {industry.title()} in {city.title()} | {date.today().strftime('%d %B %Y')}"
    title_cell.font  = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill  = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Header row ────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 22

    # Sort: HOT first, then WARM
    sorted_businesses = sorted(
        businesses,
        key=lambda b: 0 if b.get("score") == "HOT" else 1
    )

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, b in enumerate(sorted_businesses, start=3):
        score = b.get("score", "WARM")
        label = "🔴 HOT" if score == "HOT" else "🟡 WARM"

        row_data = [
            label,
            b.get("name", ""),
            b.get("category") or industry.title(),
            b.get("review_count", 0),
            b.get("phone", ""),
            b.get("website", "None"),
            b.get("address", ""),
            b.get("rating", "N/A"),
            b.get("praise", ""),
            b.get("complaints", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.row_dimensions[row_idx].height = 80

    # ── Freeze header rows ────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Save ──────────────────────────────────────────────────────────────
    os.makedirs("output", exist_ok=True)
    safe_industry = re.sub(r"[^\w\-]", "_", industry.lower())
    safe_city     = re.sub(r"[^\w\-]", "_", city.lower())
    filename = f"output/leads_{safe_industry}_{safe_city}_{date.today()}.xlsx"
    try:
        wb.save(filename)
    except OSError as e:
        raise RuntimeError(f"Failed to save Excel file '{filename}': {e}") from e
    print(f"📊 Excel saved: {filename}")

    return {"excel_path": filename}
