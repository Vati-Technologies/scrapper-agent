"""
Tests for nodes/excel_node.py
"""
import os
import pytest
from openpyxl import load_workbook
from tests.conftest import make_analyzed_business, make_pipeline_state
from nodes.excel_node import excel_node


@pytest.fixture(autouse=True)
def cleanup_output():
    """Remove generated test files after each test."""
    yield
    import glob
    for f in glob.glob("output/leads_test_*.xlsx"):
        os.remove(f)


class TestExcelNode:
    def test_creates_excel_file(self):
        state = make_pipeline_state({
            "industry":             "test_industry",
            "city":                 "test_city",
            "analyzed_businesses":  [make_analyzed_business()],
        })
        result = excel_node(state)

        assert os.path.exists(result["excel_path"])

    def test_returns_excel_path_in_state(self):
        state = make_pipeline_state({
            "industry":             "test_industry",
            "city":                 "test_city",
            "analyzed_businesses":  [make_analyzed_business()],
        })
        result = excel_node(state)

        assert "excel_path" in result
        assert result["excel_path"].endswith(".xlsx")

    def test_excel_contains_all_columns(self):
        state = make_pipeline_state({
            "industry":             "test_industry",
            "city":                 "test_city",
            "analyzed_businesses":  [make_analyzed_business()],
        })
        result = excel_node(state)
        wb = load_workbook(result["excel_path"])
        ws = wb.active

        headers = [ws.cell(row=2, column=i).value for i in range(1, 12)]
        assert "Lead Status"          in headers
        assert "Business Name"        in headers
        assert "Phone Number"         in headers
        assert "Customers Praise"     in headers
        assert "Customers Complaints" in headers
        assert "Pitch Angle"          not in headers

    def test_hot_leads_appear_before_warm(self):
        businesses = [
            make_analyzed_business({"name": "Warm Co",  "score": "WARM"}),
            make_analyzed_business({"name": "Hot Co",   "score": "HOT"}),
        ]
        state = make_pipeline_state({
            "industry":             "test_industry",
            "city":                 "test_city",
            "analyzed_businesses":  businesses,
        })
        result = excel_node(state)
        wb = load_workbook(result["excel_path"])
        ws = wb.active

        # Row 3 is first data row (row 1 = title, row 2 = headers)
        first_business_name = ws.cell(row=3, column=2).value
        assert first_business_name == "Hot Co"

    def test_excel_contains_business_data(self):
        business = make_analyzed_business({
            "name":  "Test Plumbing",
            "phone": "021 111 2222",
            "score": "HOT",
        })
        state = make_pipeline_state({
            "industry":             "test_industry",
            "city":                 "test_city",
            "analyzed_businesses":  [business],
        })
        result = excel_node(state)
        wb = load_workbook(result["excel_path"])
        ws = wb.active

        row_values = [ws.cell(row=3, column=i).value for i in range(1, 12)]
        assert "Test Plumbing" in row_values
        assert "021 111 2222"  in row_values

    def test_hot_label_shows_hot(self):
        business = make_analyzed_business({"score": "HOT"})
        state = make_pipeline_state({
            "industry":             "test_industry",
            "city":                 "test_city",
            "analyzed_businesses":  [business],
        })
        result = excel_node(state)
        wb = load_workbook(result["excel_path"])
        ws = wb.active

        status_cell = ws.cell(row=3, column=1).value
        assert "HOT" in status_cell

    def test_handles_empty_businesses(self):
        state = make_pipeline_state({
            "industry":             "test_industry",
            "city":                 "test_city",
            "analyzed_businesses":  [],
        })
        result = excel_node(state)
        # Should still create a file (just headers, no data rows)
        assert os.path.exists(result["excel_path"])
